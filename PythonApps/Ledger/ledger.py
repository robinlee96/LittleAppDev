
from datetime import datetime, timedelta
from typing import Optional, Tuple, Dict, List
from models import Account, Transaction, TransactionType
from storage import StorageManager
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


class LedgerManager:
    def __init__(self, storage: StorageManager):
        self.storage = storage

    def add_income(self, account_id: str, amount: float, category: str, description: str = "", category_id: Optional[str] = None) -> Optional[Transaction]:
        account = self.storage.get_account(account_id)
        if not account:
            return None
        if amount <= 0:
            return None

        transaction = Transaction(
            transaction_type=TransactionType.INCOME,
            amount=amount,
            category_id=category_id,
            category=category,
            description=description,
            account_id=account_id,
            created_at=datetime.now()
        )
        self.storage.add_transaction(transaction)

        new_balance = account.balance + amount
        self.storage.update_account(account_id, balance=new_balance)

        return transaction

    def add_expense(self, account_id: str, amount: float, category: str, description: str = "", category_id: Optional[str] = None) -> Optional[Transaction]:
        account = self.storage.get_account(account_id)
        if not account:
            return None
        if amount <= 0:
            return None
        if account.balance < amount:
            return None

        transaction = Transaction(
            transaction_type=TransactionType.EXPENSE,
            amount=amount,
            category_id=category_id,
            category=category,
            description=description,
            account_id=account_id,
            created_at=datetime.now()
        )
        self.storage.add_transaction(transaction)

        new_balance = account.balance - amount
        self.storage.update_account(account_id, balance=new_balance)

        return transaction

    def transfer(self, from_account_id: str, to_account_id: str, amount: float, category: str = "转账", description: str = "") -> Optional[Transaction]:
        from_account = self.storage.get_account(from_account_id)
        to_account = self.storage.get_account(to_account_id)
        if not from_account or not to_account:
            return None
        if from_account_id == to_account_id:
            return None
        if amount <= 0:
            return None
        if from_account.balance < amount:
            return None

        transaction = Transaction(
            transaction_type=TransactionType.TRANSFER,
            amount=amount,
            category=category,
            description=description,
            account_id=from_account_id,
            target_account_id=to_account_id,
            created_at=datetime.now()
        )
        self.storage.add_transaction(transaction)

        new_from_balance = from_account.balance - amount
        self.storage.update_account(from_account_id, balance=new_from_balance)

        new_to_balance = to_account.balance + amount
        self.storage.update_account(to_account_id, balance=new_to_balance)

        return transaction

    def revert_transaction(self, transaction_id: str) -> bool:
        transaction = self.storage.get_transaction(transaction_id)
        if not transaction:
            return False

        if transaction.transaction_type == TransactionType.INCOME:
            account = self.storage.get_account(transaction.account_id)
            if account:
                new_balance = account.balance - transaction.amount
                self.storage.update_account(transaction.account_id, balance=new_balance)
        elif transaction.transaction_type == TransactionType.EXPENSE:
            account = self.storage.get_account(transaction.account_id)
            if account:
                new_balance = account.balance + transaction.amount
                self.storage.update_account(transaction.account_id, balance=new_balance)
        elif transaction.transaction_type == TransactionType.TRANSFER:
            from_account = self.storage.get_account(transaction.account_id)
            to_account = self.storage.get_account(transaction.target_account_id)
            if from_account and to_account:
                new_from_balance = from_account.balance + transaction.amount
                self.storage.update_account(transaction.account_id, balance=new_from_balance)
                new_to_balance = to_account.balance - transaction.amount
                self.storage.update_account(transaction.target_account_id, balance=new_to_balance)

        self.storage.delete_transaction(transaction_id)
        return True

    def get_account_balance(self, account_id: str) -> Optional[float]:
        account = self.storage.get_account(account_id)
        if account:
            return account.balance
        return None

    def get_total_balance(self) -> float:
        return self.storage.get_total_balance()

    def get_statistics(self, period: str = "all", year: Optional[int] = None, month: Optional[int] = None) -> Dict[str, float]:
        transactions = self.storage.get_all_transactions()
        now = datetime.now()
        
        if period == "all":
            filtered = transactions
        elif period == "year" and year:
            filtered = [t for t in transactions if t.created_at.year == year]
        elif period == "month" and year and month:
            filtered = [t for t in transactions if t.created_at.year == year and t.created_at.month == month]
        elif period == "week":
            start_week = now - timedelta(days=now.weekday())
            filtered = [t for t in transactions if t.created_at.date() >= start_week.date()]
        elif period == "day":
            filtered = [t for t in transactions if t.created_at.date() == now.date()]
        else:
            filtered = []
        
        total_income = sum(t.amount for t in filtered if t.transaction_type == TransactionType.INCOME)
        total_expense = sum(t.amount for t in filtered if t.transaction_type == TransactionType.EXPENSE)
        
        return {
            "income": total_income,
            "expense": total_expense,
            "balance": total_income - total_expense
        }

    def export_to_excel(self, file_path: str, year: Optional[int] = None, month: Optional[int] = None) -> bool:
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "交易记录"
            
            headers = ["日期", "类型", "金额", "类别", "备注", "账户"]
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            transactions = self.storage.get_all_transactions()
            if year and month:
                transactions = [t for t in transactions if t.created_at.year == year and t.created_at.month == month]
            elif year:
                transactions = [t for t in transactions if t.created_at.year == year]
            
            transactions = sorted(transactions, key=lambda x: x.created_at, reverse=True)
            
            for row_num, t in enumerate(transactions, 2):
                account = self.storage.get_account(t.account_id)
                account_name = account.name if account else "-"
                if t.transaction_type == TransactionType.TRANSFER and t.target_account_id:
                    target_account = self.storage.get_account(t.target_account_id)
                    target_name = target_account.name if target_account else "-"
                    account_name = f"{account_name} -> {target_name}"
                
                ws.cell(row=row_num, column=1, value=t.created_at.strftime("%Y-%m-%d %H:%M:%S"))
                ws.cell(row=row_num, column=2, value=t.transaction_type.value)
                ws.cell(row=row_num, column=3, value=t.amount)
                ws.cell(row=row_num, column=4, value=t.category)
                ws.cell(row=row_num, column=5, value=t.description)
                ws.cell(row=row_num, column=6, value=account_name)
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(file_path)
            return True
        except Exception:
            return False
